#!/usr/bin/env python3
"""
Process S4-6 elective xlsx → basedata/student_elective_data.csv
Source: Administrative/CHW/_raw/2627_S4-6 Name List for Electives_*.xlsx
Target: chw-api/basedata/student_elective_data.csv
"""
import openpyxl, csv, os, re, glob

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RAW_DIR = os.path.join(REPO_ROOT, "Administrative", "CHW", "_raw")
CHW_API = os.path.join(os.path.dirname(REPO_ROOT), "chw-api")
OUTPUT_PATH = os.path.join(CHW_API, "basedata", "student_elective_data.csv")

INVALID = {"[object Object]", "D", "C", "object", "None", ""}


def clean_val(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s in INVALID or s.startswith("[object") or s.startswith("[Formula"):
        return ""
    return s


def process_sheet(ws, sheet_name, rows_data, correction_mode=False):
    """Extract elective data from a sheet.
    
    correction_mode: if True, only process rows after row 130 (correction rows).
    """
    start_row = 131 if correction_mode else 3
    end_row = ws.max_row + 1

    for row_idx in range(start_row, end_row):
        cls = clean_val(ws.cell(row=row_idx, column=1).value)
        num = ws.cell(row=row_idx, column=2).value
        name = clean_val(ws.cell(row=row_idx, column=3).value)
        x1 = clean_val(ws.cell(row=row_idx, column=6).value)
        x2 = clean_val(ws.cell(row=row_idx, column=7).value)
        x3 = clean_val(ws.cell(row=row_idx, column=8).value)

        if not cls or not re.match(r"^[456][A-D]$", cls):
            continue
        try:
            num_int = int(num)
        except (ValueError, TypeError):
            continue

        rows_data.append(
            {
                "class": cls,
                "number": num_int,
                "name": name,
                "X1": x1,
                "X2": x2,
                "X3": x3,
                "sheet": sheet_name,
                "is_correction": correction_mode,
            }
        )


def main():
    # Find latest xlsx in _raw
    pattern = os.path.join(RAW_DIR, "2627_S4-6 Name List for Electives*.xlsx")
    files = glob.glob(pattern)
    if not files:
        print("ERROR: No elective xlsx found in", RAW_DIR)
        return
    xlsx_path = sorted(files)[-1]
    print(f"Source: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    all_rows = []

    # Process main data from S4, S5, S6 (skip APL sheets and 總數)
    for sheet_name in ["S4", "S5", "S6"]:
        ws = wb[sheet_name]
        process_sheet(ws, sheet_name, all_rows, correction_mode=False)
        print(f"  {sheet_name}: {len([r for r in all_rows if r['sheet'] == sheet_name])} students")

    # Process correction rows (after summary section, row > 130)
    for sheet_name in ["S4", "S5"]:
        ws = wb[sheet_name]
        corrections = []
        process_sheet(ws, sheet_name, corrections, correction_mode=True)
        for corr in corrections:
            key = (corr["class"], corr["number"])
            # Remove original entry
            all_rows = [
                r
                for r in all_rows
                if not (r["class"] == key[0] and r["number"] == key[1] and r["sheet"] == sheet_name)
            ]
            all_rows.append(corr)
            print(f"  Correction applied: {key[0]} {key[1]} {corr['name']} -> {corr['X1']},{corr['X2']},{corr['X3']}")

    # Sort by class, number
    all_rows.sort(key=lambda r: (r["class"], r["number"]))

    # Write CSV
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["class", "number", "name", "X1", "X2", "X3"])
        writer.writeheader()
        for row in all_rows:
            writer.writerow(
                {
                    "class": row["class"],
                    "number": row["number"],
                    "name": row["name"],
                    "X1": row["X1"],
                    "X2": row["X2"],
                    "X3": row["X3"],
                }
            )

    # Summary
    s4 = len([r for r in all_rows if r["class"].startswith("4")])
    s5 = len([r for r in all_rows if r["class"].startswith("5")])
    s6 = len([r for r in all_rows if r["class"].startswith("6")])
    print(f"\nTotal: {len(all_rows)} records (S4:{s4}, S5:{s5}, S6:{s6})")
    print(f"Written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
