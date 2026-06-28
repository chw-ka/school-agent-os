#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Term 2 discipline + exam-absent UPDATE scripts from PY/PC Excel files.

See Administrative/CHW/student-report/guides/11_遲缺與考試缺席_Excel轉SQL.md
"""

from __future__ import annotations

import csv
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    raise

# --- paths (override with CLI args) ---
TOOL_DIR = Path(__file__).resolve().parent
DEFAULT_DISCIPLINE_XLSX = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile\2025-26_S1_S5_2nd Term_V2.xlsx"
)
DEFAULT_ABSENT_XLSX = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile\25-26_下學期考試缺席學生名單.xlsx"
)
DEFAULT_CM_AWARDS_XLSX = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\25_26_Term2_科獎及特別獎項_待CM確認.xlsx"
)
DEFAULT_MAPPING_CSV = TOOL_DIR / "data" / "exam_subject_label_to_idpaper_2526.csv"
DEFAULT_OUT_DIR = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\SQL"
)

# Discipline sheet column indices (header row 5, data from row 6) — 25-26 layout
COL_CLASS = 0
COL_NUM = 1
COL_DEMERIT_DS = 8
COL_DEMERIT_HW = 9
COL_FLG_HW_WARN = 10
COL_LATE = 15
COL_DAY_ABSENT = 16
COL_ID_STUDENT = 17

DISCIPLINE_HEADER_ROW = 5
DISCIPLINE_DATA_START = 6

# Absent primary block
ABS_COL_DATE = 0
ABS_COL_CLASS = 1
ABS_COL_NUM = 2
ABS_COL_NAME = 3
ABS_COL_ABSENT = 4
ABS_COL_LATE = 5
ABS_COL_TYPE = 6
ABS_COL_SUBJ1 = 7
ABS_COL_SUBJ2 = 8
ABS_WIDE_COL_START = 14
ABS_WIDE_COL_END = 40  # cols 41+ = 長缺日數彙總區，唔係考試卷
ABS_DATA_START_ROW = 3

EXEMPT_TYPES = {"豁免", "免考"}

# TSA 卷會入 tblStudentPaperScore，但缺席 SQL 只處理校內試，唔包括 TSA。
TSA_LABEL_MARKERS = ("(TSA)", "（TSA）")

PED_EXEMPT_SHEET = "體育豁免_參考"

# Wide-column bleed from merged cells / wrong form subjects — skip silently.
SKIP_LABELS_BY_FORM: dict[int, set[str]] = {
    3: {"化學", "物理", "生物", "生物 ", "科學"},
    2: {"化學", "物理", "生物", "生物 "},
    1: {"化學", "物理", "生物", "生物 "},
}

EXAM_KEYWORDS = re.compile(
    r"(中文|英文|Eng|Math|歷史|地理|科學|化學|物理|生物|電腦|資訊|公民|生活|經濟|"
    r"普通話|口試|Listening|Writing|Reading|Oral|GE|TSA|歷史|中國)",
    re.I,
)


def is_tsa_exam_label(label: str) -> bool:
    s = str(label).strip()
    return any(m in s for m in TSA_LABEL_MARKERS)


@dataclass
class MappingRule:
    exam_label: str
    form_min: int
    form_max: int
    id_papers: list[str]
    notes: str = ""


def load_mapping(csv_path: Path) -> list[MappingRule]:
    rules: list[MappingRule] = []
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            label = row["exam_label"].strip()
            papers = [p.strip() for p in row["idPaper"].split("|") if p.strip()]
            rules.append(
                MappingRule(
                    exam_label=label,
                    form_min=int(row["form_min"]),
                    form_max=int(row["form_max"]),
                    id_papers=papers,
                    notes=row.get("notes", ""),
                )
            )
    return rules


def normalize_label(label: str) -> str:
    s = str(label).strip()
    s = re.sub(r"^\([0-9][A-D]\)\s*", "", s)
    return s


def form_from_class(class_name: str) -> int | None:
    if not class_name:
        return None
    m = re.match(r"^(\d)", str(class_name).strip())
    return int(m.group(1)) if m else None


def looks_like_student_name(label: str) -> bool:
    s = str(label).strip()
    if not s or EXAM_KEYWORDS.search(s):
        return False
    return bool(re.fullmatch(r"[\u4e00-\u9fff]{2,4}", s))


def should_skip_label(label: str, form: int) -> bool:
    norm = normalize_label(label)
    if looks_like_student_name(norm):
        return True
    skip = SKIP_LABELS_BY_FORM.get(form, set())
    return norm in skip or label.strip() in skip


def resolve_id_papers(
    raw_label: str, form: int, rules: list[MappingRule]
) -> tuple[list[str], str | None]:
    label = normalize_label(raw_label)
    if not label:
        return [], None
    matches = [
        r
        for r in rules
        if r.exam_label == label or normalize_label(r.exam_label) == label
    ]
    if not matches:
        # fuzzy: ignore parenthetical class prefix in CSV
        for r in rules:
            if normalize_label(r.exam_label) == label:
                matches.append(r)
    if not matches:
        return [], label
    for r in matches:
        if r.form_min <= form <= r.form_max:
            return r.id_papers, None
    return [], f"{label} (form {form} out of range)"


def resolve_papers(
    raw_label: str, form: int, rules: list[MappingRule]
) -> tuple[list[str], str | None]:
    if should_skip_label(raw_label, form):
        return [], None
    return resolve_id_papers(raw_label, form, rules)


def to_int(value, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def to_float(value, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def to_bit_y(value) -> int:
    if value is None:
        return 0
    s = str(value).strip().upper()
    return 1 if s in ("Y", "1", "TRUE", "YES") else 0


def read_discipline_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if "缺點" in name or "20260622" in name:
            sheet_name = name
            break
    ws = wb[sheet_name or wb.sheetnames[0]]
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < DISCIPLINE_DATA_START:
            continue
        if not row[COL_CLASS] or not row[COL_ID_STUDENT]:
            continue
        rows.append(
            {
                "class": str(row[COL_CLASS]).strip(),
                "num": row[COL_NUM],
                "id_student": int(float(row[COL_ID_STUDENT])),
                "day_absent": to_float(row[COL_DAY_ABSENT]),
                "num_late": to_int(row[COL_LATE]),
                "demerit_ds": to_int(row[COL_DEMERIT_DS]),
                "demerit_hw": to_int(row[COL_DEMERIT_HW]),
                "flg_hw": to_bit_y(row[COL_FLG_HW_WARN]),
            }
        )
    wb.close()
    return rows


def generate_discipline_sql(rows: list[dict]) -> str:
    lines = [
        "-- Generated: Update tblStudentDiscipline term 2 (遲缺、缺點、功課警告)",
        f"-- Rows: {len(rows)}",
        f"-- {datetime.now().isoformat(timespec='seconds')}",
        "USE db25_26;",
        "GO",
        "",
    ]
    for r in rows:
        lines.append(
            "UPDATE tblStudentDiscipline SET "
            f"dayAbsent_2 = {r['day_absent']}, "
            f"numLate_2 = {r['num_late']}, "
            f"numDemeritDS_2 = {r['demerit_ds']}, "
            f"numDemeritHW_2 = {r['demerit_hw']}, "
            f"flgHW_2 = {r['flg_hw']} "
            f"WHERE idStudent = {r['id_student']};"
        )
    lines.append("")
    lines.append(f"-- Done: {len(rows)} updates")
    return "\n".join(lines)


@dataclass
class AbsentRecord:
    id_student: int | None
    class_name: str
    num: int | float
    name: str
    exam_label: str
    absent_type: str
    is_exempt: bool
    source: str


def lookup_id_student(
    class_name: str, num: int | float, student_map: dict[tuple[str, int], int]
) -> int | None:
    key = (str(class_name).strip(), int(float(num)))
    return student_map.get(key)


def build_student_map(discipline_rows: list[dict]) -> dict[tuple[str, int], int]:
    m: dict[tuple[str, int], int] = {}
    for r in discipline_rows:
        m[(r["class"], int(float(r["num"])))] = r["id_student"]
    return m


def read_absent_records(
    xlsx_path: Path, student_map: dict[tuple[str, int], int]
) -> tuple[list[AbsentRecord], list[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records: list[AbsentRecord] = []
    unmapped: list[str] = []
    seen: set[tuple] = set()

    def add_record(
        class_name: str,
        num,
        name: str,
        label,
        absent_type: str,
        source: str,
    ) -> None:
        if not label or not class_name or num is None:
            return
        label_s = str(label).strip()
        if not label_s or label_s in ("沒有考/遲到", "absent", "late"):
            return
        is_exempt = str(absent_type).strip() in EXEMPT_TYPES
        key = (class_name, int(float(num)), label_s, is_exempt)
        if key in seen:
            return
        seen.add(key)
        sid = lookup_id_student(class_name, num, student_map)
        records.append(
            AbsentRecord(
                id_student=sid,
                class_name=str(class_name).strip(),
                num=num,
                name=str(name or "").strip(),
                exam_label=label_s,
                absent_type=str(absent_type or "").strip(),
                is_exempt=is_exempt,
                source=source,
            )
        )

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < ABS_DATA_START_ROW:
            continue
        if not row[ABS_COL_CLASS]:
            continue

        # Primary block: only when absent flag set
        if row[ABS_COL_ABSENT] == 1:
            cls, num, name = row[ABS_COL_CLASS], row[ABS_COL_NUM], row[ABS_COL_NAME]
            atype = row[ABS_COL_TYPE] or ""
            add_record(cls, num, name, row[ABS_COL_SUBJ1], atype, f"row{i}:subj1")
            add_record(cls, num, name, row[ABS_COL_SUBJ2], atype, f"row{i}:subj2")

        # Wide columns: only for 長缺 — lists all missed exam papers
        if str(row[ABS_COL_TYPE] or "").strip() == "長缺":
            cls, num, name = row[ABS_COL_CLASS], row[ABS_COL_NUM], row[ABS_COL_NAME]
            atype = "長缺"
            for c in range(ABS_WIDE_COL_START, ABS_WIDE_COL_END):
                val = row[c] if c < len(row) else None
                if isinstance(val, str) and val.strip():
                    add_record(cls, num, name, val, atype, f"row{i}:wide{c}")

    wb.close()
    return records, unmapped


def read_ped_exempt_records(
    xlsx_path: Path, student_map: dict[tuple[str, int], int]
) -> list[AbsentRecord]:
    if not xlsx_path.exists():
        return []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if PED_EXEMPT_SHEET not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[PED_EXEMPT_SHEET]
    records: list[AbsentRecord] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 4 or not row[2]:
            continue
        class_name = str(row[2]).strip()
        num = row[3]
        sid = row[5] if len(row) > 5 and row[5] else None
        if sid is None:
            sid = lookup_id_student(class_name, num, student_map)
        records.append(
            AbsentRecord(
                id_student=int(sid) if sid is not None else None,
                class_name=class_name,
                num=num,
                name=str(row[4] or "").strip(),
                exam_label="PED",
                absent_type="體育豁免",
                is_exempt=True,
                source=f"體育豁免 row{i}",
            )
        )
    wb.close()
    return records


@dataclass
class PaperUpdate:
    id_student: int
    id_paper: str
    is_exempt: bool
    comment: str


def collect_paper_updates(
    records: list[AbsentRecord],
    rules: list[MappingRule],
    student_map: dict[tuple[str, int], int],
) -> tuple[list[PaperUpdate], list[str], list[str], list[str]]:
    """Build deduped (idStudent, idPaper) updates; matches Excel generator logic."""
    unmapped: list[str] = []
    skipped_tsa: list[str] = []
    missing_id: list[str] = []
    updates: dict[tuple[int, str, bool], PaperUpdate] = {}

    def add_update(sid: int, paper: str, is_exempt: bool, comment: str) -> None:
        key = (sid, paper, is_exempt)
        if key not in updates:
            updates[key] = PaperUpdate(sid, paper, is_exempt, comment)

    for rec in records:
        if rec.is_exempt and rec.exam_label == "PED":
            sid = rec.id_student
            if sid is None:
                sid = lookup_id_student(rec.class_name, rec.num, student_map)
            if sid is None:
                missing_id.append(
                    f"{rec.class_name}{rec.num} {rec.name} PED ({rec.source})"
                )
                continue
            add_update(
                sid,
                "PED",
                True,
                f"{rec.class_name} {rec.num} {rec.name} | PED | {rec.absent_type}",
            )
            continue

        if is_tsa_exam_label(rec.exam_label):
            skipped_tsa.append(
                f"{rec.class_name}{rec.num} {rec.name} '{rec.exam_label}' ({rec.absent_type})"
            )
            continue
        form = form_from_class(rec.class_name)
        if form is None:
            unmapped.append(f"{rec.class_name}{rec.num} {rec.exam_label}: unknown form")
            continue
        papers, err = resolve_papers(rec.exam_label, form, rules)
        if err:
            unmapped.append(f"{rec.class_name}{rec.num} {rec.name}: {err}")
            continue
        if not papers:
            continue

        sid = rec.id_student
        if sid is None:
            sid = lookup_id_student(rec.class_name, rec.num, student_map)
        if sid is None:
            missing_id.append(
                f"{rec.class_name}{rec.num} {rec.name} '{rec.exam_label}' ({rec.source})"
            )
            continue

        for paper in papers:
            add_update(
                sid,
                paper,
                rec.is_exempt,
                (
                    f"{rec.class_name} {rec.num} {rec.name} | "
                    f"{rec.exam_label} -> {paper} | {rec.absent_type}"
                ),
            )

    ordered = sorted(updates.values(), key=lambda u: (u.id_student, u.id_paper, u.is_exempt))
    return ordered, unmapped, missing_id, skipped_tsa


def generate_absent_sql(
    records: list[AbsentRecord],
    rules: list[MappingRule],
    student_map: dict[tuple[str, int], int],
) -> tuple[str, list[str], list[str]]:
    paper_updates, unmapped, missing_id, skipped_tsa = collect_paper_updates(
        records, rules, student_map
    )
    lines = [
        "-- Generated: Exam absent / exempt for term 2",
        f"-- Source records: {len(records)}",
        f"-- Unique paper updates: {len(paper_updates)}",
        f"-- {datetime.now().isoformat(timespec='seconds')}",
        "USE db25_26;",
        "GO",
        "",
    ]

    for upd in paper_updates:
        flag_sql = "flgIgnore_2 = 1" if upd.is_exempt else "flgAbsent_2 = 1"
        lines.append(f"-- {upd.comment}")
        lines.append(
            "UPDATE tblStudentPaperScore SET "
            f"score_exam_2 = 0, {flag_sql} "
            f"WHERE idStudent = {upd.id_student} AND idPaper = '{upd.id_paper}';"
        )

    absent_n = sum(1 for u in paper_updates if not u.is_exempt)
    exempt_n = sum(1 for u in paper_updates if u.is_exempt)
    lines.append("")
    lines.append(
        f"-- Done: {len(paper_updates)} paper-level updates "
        f"(absent {absent_n}, exempt {exempt_n})"
    )
    if skipped_tsa:
        lines.append(f"-- Skipped TSA (not 校內試): {len(skipped_tsa)} source records")
    if missing_id:
        lines.append("-- WARNING: could not resolve idStudent for:")
        for m in missing_id:
            lines.append(f"--   {m}")
    return "\n".join(lines), unmapped, skipped_tsa


def main() -> int:
    discipline_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_DISCIPLINE_XLSX
    absent_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_ABSENT_XLSX
    mapping_path = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_MAPPING_CSV
    out_dir = Path(sys.argv[4]) if len(sys.argv) > 4 else DEFAULT_OUT_DIR
    cm_awards_path = Path(sys.argv[5]) if len(sys.argv) > 5 else DEFAULT_CM_AWARDS_XLSX

    if not discipline_path.exists():
        print(f"Missing discipline file: {discipline_path}", file=sys.stderr)
        return 1
    if not absent_path.exists():
        print(f"Missing absent file: {absent_path}", file=sys.stderr)
        return 1
    if not mapping_path.exists():
        print(f"Missing mapping CSV: {mapping_path}", file=sys.stderr)
        return 1

    out_dir.mkdir(parents=True, exist_ok=True)

    discipline_rows = read_discipline_rows(discipline_path)
    discipline_sql = generate_discipline_sql(discipline_rows)
    disc_out = out_dir / "01_Update_tblStudentDiscipline_term2.sql"
    disc_out.write_text(discipline_sql, encoding="utf-8")
    print(f"Wrote {disc_out} ({len(discipline_rows)} students)")

    rules = load_mapping(mapping_path)
    student_map = build_student_map(discipline_rows)
    absent_records, _ = read_absent_records(absent_path, student_map)
    ped_records = read_ped_exempt_records(cm_awards_path, student_map)
    all_records = absent_records + ped_records
    absent_sql, issues, skipped_tsa = generate_absent_sql(all_records, rules, student_map)
    abs_out = out_dir / "02_Update_tblStudentPaperScore_exam_absent_term2.sql"
    abs_out.write_text(absent_sql, encoding="utf-8")
    print(
        f"Wrote {abs_out} ({len(absent_records)} absent + {len(ped_records)} PED exempt parsed)"
    )
    if skipped_tsa:
        tsa_out = out_dir / "02_skipped_tsa_absent_records.txt"
        tsa_out.write_text("\n".join(skipped_tsa), encoding="utf-8")
        print(f"Skipped {len(skipped_tsa)} TSA records -> {tsa_out}")

    if issues:
        issue_out = out_dir / "02_unmapped_exam_labels.txt"
        issue_out.write_text("\n".join(sorted(set(issues))), encoding="utf-8")
        print(f"Wrote {issue_out} ({len(set(issues))} items need review)")
    else:
        print("All exam labels mapped.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
