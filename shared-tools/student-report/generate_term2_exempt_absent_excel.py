#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build A_Done_SQL_Term2_Exempt_Absent.xlsx for 25-26 from last year's template.

Copies the 24-25 workbook shell (formulas / sheet layout), refreshes StudentList,
and fills 豁免_缺席_Term2 from:
  - 25-26_下學期考試缺席學生名單.xlsx
  - 25_26_Term2_科獎及特別獎項_待CM確認.xlsx (體育豁免_參考)

See Administrative/CHW/student-report/guides/11_遲缺與考試缺席_Excel轉SQL.md
"""

from __future__ import annotations

import shutil
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# Reuse parsers / mapping from SQL generator
from generate_term2_discipline_absent_sql import (
    DEFAULT_ABSENT_XLSX,
    DEFAULT_DISCIPLINE_XLSX,
    DEFAULT_MAPPING_CSV,
    AbsentRecord,
    build_student_map,
    form_from_class,
    is_tsa_exam_label,
    load_mapping,
    read_absent_records,
    read_ped_exempt_records,
    resolve_papers,
)

TEMPLATE_XLSX = Path(
    r"T:\24-25\ITAdmin_13_StudentReport\Datafile\24_25_Term2"
    r"\A_Done_SQL_Term2_Exempt_Absent.xlsx"
)
DEFAULT_CM_AWARDS_XLSX = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\25_26_Term2_科獎及特別獎項_待CM確認.xlsx"
)
DEFAULT_OUT_XLSX = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2"
    r"\A_Done_SQL_Term2_Exempt_Absent.xlsx"
)

MAIN_SHEET_HINT = "缺席"
STUDENT_LIST_SHEET = "StudentList"
PED_EXEMPT_SHEET = "體育豁免_參考"


@dataclass
class OutputRow:
    class_name: str
    num: int
    id_paper: str
    absent: bool
    exempt: bool
    source: str


def records_to_rows(records: list[AbsentRecord], rules) -> tuple[list[OutputRow], list[str]]:
    rows: list[OutputRow] = []
    issues: list[str] = []
    seen: set[tuple] = set()

    for rec in records:
        if is_tsa_exam_label(rec.exam_label):
            continue
        form = form_from_class(rec.class_name)
        if form is None:
            issues.append(f"{rec.class_name}{rec.num}: unknown form for '{rec.exam_label}'")
            continue
        papers, err = resolve_papers(rec.exam_label, form, rules)
        if err:
            issues.append(f"{rec.class_name}{rec.num} {rec.name}: {err}")
            continue
        if not papers:
            continue
        for paper in papers:
            key = (rec.class_name, int(float(rec.num)), paper, rec.is_exempt)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                OutputRow(
                    class_name=rec.class_name,
                    num=int(float(rec.num)),
                    id_paper=paper,
                    absent=not rec.is_exempt,
                    exempt=rec.is_exempt,
                    source=rec.source,
                )
            )
    return rows, issues


def read_ped_exempt_rows(
    cm_xlsx: Path, student_map: dict[tuple[str, int], int]
) -> list[OutputRow]:
    records = read_ped_exempt_records(cm_xlsx, student_map)
    rows: list[OutputRow] = []
    for rec in records:
        rows.append(
            OutputRow(
                class_name=rec.class_name,
                num=int(float(rec.num)),
                id_paper="PED",
                absent=False,
                exempt=True,
                source=rec.source,
            )
        )
    return rows


def merge_rows(absent_rows: list[OutputRow], ped_rows: list[OutputRow]) -> list[OutputRow]:
    merged: list[OutputRow] = []
    seen: set[tuple] = set()
    for row in absent_rows + ped_rows:
        key = (row.class_name, row.num, row.id_paper, row.exempt)
        if key in seen:
            continue
        seen.add(key)
        merged.append(row)
    merged.sort(key=lambda r: (r.class_name, r.num, r.id_paper))
    return merged


def read_discipline_student_list(xlsx_path: Path) -> list[dict]:
    """Class/num/idStudent/name rows for StudentList sheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if "缺點" in name or "20260622" in name:
            sheet_name = name
            break
    ws = wb[sheet_name or wb.sheetnames[0]]
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 6 or not row[0] or not row[17]:
            continue
        rows.append(
            {
                "class": str(row[0]).strip(),
                "num": row[1],
                "name": str(row[2]).strip() if row[2] else "",
                "id_student": int(float(row[17])),
            }
        )
    wb.close()
    return rows


def write_student_list(ws, discipline_rows: list[dict]) -> None:
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r in discipline_rows:
        cls = r["class"]
        num = int(float(r["num"]))
        ws.append([f"{cls}{num}", r["id_student"], cls, num, r.get("name") or None])


def set_main_formulas(ws, row: int) -> None:
    ws.cell(row, 1).value = (
        f'="(idStudent = " & C{row} & " and idPaper = \'" & G{row} & "\')"'
    )
    if row == 3:
        ws.cell(row, 2).value = f"=A{row}"
    else:
        ws.cell(row, 2).value = f'=B{row - 1}& " or " & A{row}'
    ws.cell(row, 3).value = (
        f'=IF(ISNA(VLOOKUP(E{row}&F{row},StudentList!A:E,2,FALSE)),"",'
        f"VLOOKUP(E{row}&F{row},StudentList!A:E,2,FALSE))"
    )
    ws.cell(row, 4).value = (
        f'=IF(ISNA(VLOOKUP(E{row}&F{row},StudentList!A:E,5,FALSE)),"",'
        f"VLOOKUP(E{row}&F{row},StudentList!A:E,5,FALSE))"
    )
    ws.cell(row, 10).value = (
        f'=IF(C{row}<>"","update tblStudentPaperScore set score_exam_2 = 0" & '
        f'IF(H{row}=1, ", flgAbsent_2 = 1", IF(I{row}=1, ", flgIgnore_2=1","")) & '
        f'" where idStudent = " & C{row} & " and idPaper = \'" & G{row} &"\'","")'
    )


def clear_main_data(ws, from_row: int) -> None:
    for row in range(from_row, ws.max_row + 1):
        for col in range(1, 11):
            ws.cell(row, col).value = None


def write_main_sheet(ws, rows: list[OutputRow]) -> None:
    clear_main_data(ws, 3)
    for i, data in enumerate(rows, start=3):
        set_main_formulas(ws, i)
        ws.cell(i, 5).value = data.class_name
        ws.cell(i, 6).value = data.num
        ws.cell(i, 7).value = data.id_paper
        ws.cell(i, 8).value = 1 if data.absent else None
        ws.cell(i, 9).value = 1 if data.exempt else None


def find_main_sheet(wb: openpyxl.Workbook):
    for name in wb.sheetnames:
        if MAIN_SHEET_HINT in name:
            return wb[name]
    return wb[wb.sheetnames[0]]


def generate_excel(
    template: Path,
    out_xlsx: Path,
    discipline_xlsx: Path,
    absent_xlsx: Path,
    cm_awards_xlsx: Path,
    mapping_csv: Path,
) -> tuple[int, list[str]]:
    if not template.exists():
        raise FileNotFoundError(f"Template missing: {template}")
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, out_xlsx)

    discipline_rows = read_discipline_student_list(discipline_xlsx)
    rules = load_mapping(mapping_csv)
    student_map = build_student_map(
        [{"class": r["class"], "num": r["num"], "id_student": r["id_student"]} for r in discipline_rows]
    )
    absent_records, _ = read_absent_records(absent_xlsx, student_map)
    absent_rows, issues = records_to_rows(absent_records, rules)
    ped_rows = read_ped_exempt_rows(cm_awards_xlsx, student_map)
    all_rows = merge_rows(absent_rows, ped_rows)

    wb = openpyxl.load_workbook(out_xlsx)
    main_ws = find_main_sheet(wb)
    sl_ws = wb[STUDENT_LIST_SHEET]

    write_student_list(sl_ws, discipline_rows)
    write_main_sheet(main_ws, all_rows)

    # Update SQL sheet note (keep query text; refresh label only)
    if "SQL" in wb.sheetnames:
        sql_ws = wb["SQL"]
        sql_ws["A2"] = "刷新 StudentList"
        sql_ws["A3"] = (
            "select class + cast(numberClass as varchar), idStudent, class, "
            "numberClass, nameChinese"
        )
        sql_ws["A4"] = "from tblStudent"
        sql_ws["A5"] = "order by class, numberClass"

    wb.save(out_xlsx)
    wb.close()
    return len(all_rows), issues


def main() -> int:
    template = Path(sys.argv[1]) if len(sys.argv) > 1 else TEMPLATE_XLSX
    out_xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT_XLSX
    discipline_xlsx = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_DISCIPLINE_XLSX
    absent_xlsx = Path(sys.argv[4]) if len(sys.argv) > 4 else DEFAULT_ABSENT_XLSX
    cm_awards_xlsx = Path(sys.argv[5]) if len(sys.argv) > 5 else DEFAULT_CM_AWARDS_XLSX
    mapping_csv = Path(sys.argv[6]) if len(sys.argv) > 6 else DEFAULT_MAPPING_CSV

    for label, path in [
        ("discipline", discipline_xlsx),
        ("absent", absent_xlsx),
        ("cm awards", cm_awards_xlsx),
        ("mapping", mapping_csv),
    ]:
        if not path.exists():
            print(f"Missing {label} file: {path}", file=sys.stderr)
            return 1

    count, issues = generate_excel(
        template, out_xlsx, discipline_xlsx, absent_xlsx, cm_awards_xlsx, mapping_csv
    )
    print(f"Wrote {out_xlsx} ({count} data rows)")

    issue_path = out_xlsx.parent / "A_Done_SQL_Term2_Exempt_Absent_unmapped.txt"
    if issues:
        issue_path.write_text("\n".join(sorted(set(issues))), encoding="utf-8")
        print(f"Review {len(set(issues))} unmapped labels -> {issue_path}")
    elif issue_path.exists():
        issue_path.unlink()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
