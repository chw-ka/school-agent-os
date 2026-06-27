#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate D_Done_SQL_Term2_Merit.xlsx by copying 24-25 template and filling from ES merit source.

Preserves colleague Excel format / I-column INSERT formula; only A-H data rows change.
"""

from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    import win32com.client as win32
except ImportError:
    print("Install: pip install openpyxl pywin32", file=sys.stderr)
    raise

DEFAULT_TEMPLATE = Path(
    r"T:\24-25\ITAdmin_13_StudentReport\Datafile\24_25_Term2\D_Done_SQL_Term2_Merit.xlsx"
)
DEFAULT_SOURCE = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\1_優點記錄(2025-2026全年)_申請及批核_S1_5__FINAL_ZL.xlsx"
)
DEFAULT_OUT = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\D_Done_SQL_Term2_Merit.xlsx"
)

DATA_START_ROW = 2
SOURCE_SHEET_HINT = "2025-2026(S1-5)"
# Source F:I (優點1-4) -> output E:H
MERIT_SRC_COLS = (6, 7, 8, 9)
MERIT_DST_COLS = (5, 6, 7, 8)
SQL_FORMULA_COL = 9


def merit_value(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return int(value)


def read_source_rows(xlsx_path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if SOURCE_SHEET_HINT in name or name.startswith("2025-2026"):
            sheet_name = name
            break
    if not sheet_name:
        raise ValueError(f"No source sheet matching {SOURCE_SHEET_HINT!r} in {xlsx_path}")
    ws = wb[sheet_name]
    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < DATA_START_ROW:
            continue
        if not row[0]:
            continue
        rows.append(tuple(row))
    wb.close()
    return rows


def ensure_data_rows(ws, xl, n_students: int, current_last_row: int) -> int:
    last_needed = DATA_START_ROW + n_students - 1
    if last_needed > current_last_row:
        ws.Range(f"{DATA_START_ROW}:{DATA_START_ROW}").Copy()
        ws.Range(f"{current_last_row + 1}:{last_needed}").Insert()
        xl.CutCopyMode = False
        current_last_row = last_needed
    elif last_needed < current_last_row:
        ws.Range(
            ws.Cells(last_needed + 1, 1),
            ws.Cells(current_last_row, SQL_FORMULA_COL),
        ).ClearContents()
        current_last_row = last_needed
    return current_last_row


def populate_excel(template_path: Path, out_path: Path, source_path: Path) -> int:
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    if not source_path.exists():
        raise FileNotFoundError(f"Source not found: {source_path}")

    source_rows = read_source_rows(source_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, out_path)

    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = None
    try:
        wb = xl.Workbooks.Open(str(out_path.resolve()))
        ws = wb.Sheets(1)
        current_last = ws.UsedRange.Rows.Count
        current_last = ensure_data_rows(ws, xl, len(source_rows), current_last)

        for offset, row in enumerate(source_rows):
            r = DATA_START_ROW + offset
            try:
                ws.Range(ws.Cells(r, 1), ws.Cells(r, 4)).Value = (
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                )
                merits = tuple(
                    merit_value(row[src_col - 1] if src_col - 1 < len(row) else None)
                    for src_col in MERIT_SRC_COLS
                )
                ws.Range(ws.Cells(r, 5), ws.Cells(r, 8)).Value = merits
            except Exception as exc:
                raise RuntimeError(f"Failed writing Excel row {r} (source offset {offset})") from exc

        wb.Save()
        return len(source_rows)
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=True)
            except Exception:
                pass
        try:
            xl.Quit()
        except Exception:
            pass


def main() -> int:
    template_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_TEMPLATE
    source_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_SOURCE
    out_path = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_OUT

    n = populate_excel(template_path, out_path, source_path)
    print(f"Wrote {out_path} ({n} students, {datetime.now().isoformat(timespec='seconds')})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
