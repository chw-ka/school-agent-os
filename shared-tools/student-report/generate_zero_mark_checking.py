#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Term 2 zero / missing score checking Excel (B_ZeroMarkChecking.xlsx).

Mirrors 24-25 B_Done_ZeroMarkChecking.xlsx layout and message templates.
Data source: db25_26, same logic as Check Record Entries (term 2).sql.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path

try:
    import openpyxl
    import pyodbc
except ImportError:
    print("Install: pip install openpyxl pyodbc", file=sys.stderr)
    raise

from _mssql_conn import connection_string

CONN = connection_string()

DEFAULT_OUT = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\B_ZeroMarkChecking.xlsx"
)

TEMPLATE = Path(
    r"T:\24-25\ITAdmin_13_StudentReport\Datafile\24_25_Term2\B_Done_ZeroMarkChecking.xlsx"
)

SCORE_SQL = """
SELECT DISTINCT
    ss.idStaff,
    sp.class AS auxiliaryClass,
    sps.idPaper,
    s.idStudent,
    s.class,
    s.numberClass,
    s.nameChinese,
    s.gender,
    sps.score_test_2,
    sps.score_regular_2,
    sps.score_exam_2,
    sps.grade_exam_2,
    fps.score_test_2 AS fm_test,
    fps.score_regular_2 AS fm_regular,
    fps.score_exam_2 AS fm_exam
FROM dbo.tblStudent s
INNER JOIN dbo.tblStudentPaperScore sps ON s.idStudent = sps.idStudent
INNER JOIN dbo.vwStudentPaper sp
    ON sps.idStudent = sp.idStudent AND sps.idPaper = sp.idPaper
INNER JOIN dbo.tblFormPaperScore fps
    ON sp.form = fps.form AND sp.idPaper = fps.idPaper
INNER JOIN dbo.tblPaper p
    ON sp.idPaper = sps.idPaper AND p.formGroup = sp.formGroup
INNER JOIN dbo.vwStaffSubject ss
    ON sp.idSubject = ss.idSubject AND sp.class = ss.class AND ss.flgTeach = 1
WHERE sp.form BETWEEN 1 AND 5
  AND sp.flgTerm2 = 1
  AND sps.flgIgnore_2 = 0
  AND sps.flgAbsent_2 = 0
  AND sps.idPaper NOT IN (N'PED', N'MUS')
  AND (
        (fps.score_test_2 IS NOT NULL AND (sps.score_test_2 IS NULL OR sps.score_test_2 = 0))
     OR (fps.score_regular_2 IS NOT NULL AND (sps.score_regular_2 IS NULL OR sps.score_regular_2 = 0))
     OR (fps.score_exam_2 IS NOT NULL AND (sps.score_exam_2 IS NULL OR sps.score_exam_2 = 0))
      )
ORDER BY ss.idStaff, sp.class, sps.idPaper, s.class, s.numberClass
"""

GRADE_SQL = """
SELECT DISTINCT
    ss.idStaff,
    sp.class AS auxiliaryClass,
    sps.idPaper,
    s.idStudent,
    s.class,
    s.numberClass,
    s.nameChinese,
    s.gender
FROM dbo.tblStudent s
INNER JOIN dbo.tblStudentPaperScore sps ON s.idStudent = sps.idStudent
INNER JOIN dbo.vwStudentPaper sp
    ON sps.idStudent = sp.idStudent AND sps.idPaper = sp.idPaper
INNER JOIN dbo.tblPaper p
    ON sp.idPaper = sps.idPaper AND p.formGroup = sp.formGroup
INNER JOIN dbo.vwStaffSubject ss
    ON sp.idSubject = ss.idSubject AND sp.class = ss.class AND ss.flgTeach = 1
WHERE sps.idPaper IN (N'PED', N'MUS')
  AND sps.grade_exam_2 IS NULL
  AND sp.form BETWEEN 1 AND 5
  AND sp.flgTerm2 = 1
  AND sps.flgIgnore_2 = 0
  AND sps.flgAbsent_2 = 0
ORDER BY ss.idStaff, sp.class, sps.idPaper, s.class, s.numberClass
"""

COMP_CN = {"test": "測驗", "regular": "平時分", "exam": "考試分"}


@dataclass
class ScoreRow:
    idStaff: str
    auxiliaryClass: str
    idPaper: str
    idStudent: int
    class_: str
    numberClass: int
    nameChinese: str
    gender: str
    score_test_2: object
    score_regular_2: object
    score_exam_2: object
    fm_test: object
    fm_regular: object
    fm_exam: object


def _state(score) -> str:
    return "NULL" if score is None else "0"


def _disp(score):
    return "NULL" if score is None else score


def _label(paper: str, comp: str, state: str) -> str:
    return f"{paper}{COMP_CN[comp]}是{state}"


def _student_ref(row: ScoreRow) -> str:
    num = f"{row.numberClass:02d}" if row.numberClass is not None else ""
    return f"{row.class_}{num}{row.nameChinese}"


def _issues(row: ScoreRow) -> dict[str, str]:
    out: dict[str, str] = {}
    if row.fm_test is not None and (row.score_test_2 is None or row.score_test_2 == 0):
        out["test"] = _state(row.score_test_2)
    if row.fm_regular is not None and (row.score_regular_2 is None or row.score_regular_2 == 0):
        out["regular"] = _state(row.score_regular_2)
    if row.fm_exam is not None and (row.score_exam_2 is None or row.score_exam_2 == 0):
        out["exam"] = _state(row.score_exam_2)
    return out


def _expand_score_rows(rows: list[ScoreRow]) -> list[dict]:
    """One output row per issue; combine regular+exam when both are NULL."""
    expanded: list[dict] = []
    for row in rows:
        issues = _issues(row)
        if not issues:
            continue

        reg_null = issues.get("regular") == "NULL"
        exam_null = issues.get("exam") == "NULL"
        combined = reg_null and exam_null

        if combined:
            reg_lbl = _label(row.idPaper, "regular", "NULL")
            exam_lbl = _label(row.idPaper, "exam", "NULL")
            combined_lbl = f"{reg_lbl}；{exam_lbl}"
            msg = (
                f"{{n}}. {row.idStaff}你好，想與你確認一下 {_student_ref(row)} 的 "
                f"{combined_lbl}是否正確？還是缺席考試？"
            )
            expanded.append(
                {
                    "staff": row.idStaff,
                    "aux": row.auxiliaryClass,
                    "paper": row.idPaper,
                    "idStudent": row.idStudent,
                    "class": row.class_,
                    "num": row.numberClass,
                    "name": row.nameChinese,
                    "gender": row.gender,
                    "c11": "NULL",
                    "c12": "NULL",
                    "c13": "NULL",
                    "c15": 1,
                    "c16": 1,
                    "c18": reg_lbl,
                    "c19": exam_lbl,
                    "c20": combined_lbl,
                    "msg": msg,
                }
            )
            issues.pop("regular", None)
            issues.pop("exam", None)

        for comp in ("test", "regular", "exam"):
            if comp not in issues:
                continue
            state = issues[comp]
            lbl = _label(row.idPaper, comp, state)
            c11 = _disp(row.score_test_2) if comp == "test" else "NULL"
            c12 = _disp(row.score_regular_2) if comp == "regular" else "NULL"
            c13 = _disp(row.score_exam_2) if comp == "exam" else "NULL"
            c15 = 1 if comp == "regular" else None
            c16 = 1 if comp == "exam" else None
            c18 = lbl if comp == "regular" else None
            c19 = lbl if comp == "exam" else None
            if comp == "test":
                c18 = lbl
                c19 = None
            if comp in ("exam", "test") and state == "NULL":
                tail = "是否正確？還是缺席考試？"
            elif comp == "exam" and state == "0":
                tail = "是否正確？還是缺席考試？"
            else:
                tail = "是否正確？"
            msg = (
                f"{{n}}. {row.idStaff}你好，想與你確認一下 {_student_ref(row)} 的 {lbl}{tail}"
            )
            expanded.append(
                {
                    "staff": row.idStaff,
                    "aux": row.auxiliaryClass,
                    "paper": row.idPaper,
                    "idStudent": row.idStudent,
                    "class": row.class_,
                    "num": row.numberClass,
                    "name": row.nameChinese,
                    "gender": row.gender,
                    "c11": c11,
                    "c12": c12,
                    "c13": c13,
                    "c15": c15,
                    "c16": c16,
                    "c18": c18,
                    "c19": c19,
                    "c20": lbl,
                    "msg": msg,
                }
            )
    return expanded


def _expand_grade_rows(rows: list[dict]) -> list[dict]:
    expanded: list[dict] = []
    for row in rows:
        paper = row["idPaper"]
        lbl = f"{paper}考試分是NULL"
        ref = f"{row['class']}{row['numberClass']:02d}{row['nameChinese']}"
        msg = (
            f"{{n}}. {row['idStaff']}你好，想與你確認一下 {ref} 的 {lbl}"
            "是否正確？還是缺席考試？"
        )
        expanded.append(
            {
                "staff": row["idStaff"],
                "aux": row["auxiliaryClass"],
                "paper": paper,
                "idStudent": row["idStudent"],
                "class": row["class"],
                "num": row["numberClass"],
                "name": row["nameChinese"],
                "gender": row["gender"],
                "c11": "NULL",
                "c12": "NULL",
                "c13": "NULL",
                "c15": None,
                "c16": 1,
                "c18": None,
                "c19": lbl,
                "c20": lbl,
                "msg": msg,
            }
        )
    return expanded


def _assign_seq(rows: list[dict]) -> None:
    counts: dict[str, int] = {}
    for r in rows:
        counts[r["staff"]] = counts.get(r["staff"], 0) + 1
        r["seq"] = counts[r["staff"]]
        r["msg"] = r["msg"].format(n=r["seq"])


def _load_workbook() -> openpyxl.Workbook:
    if TEMPLATE.exists():
        wb = openpyxl.load_workbook(TEMPLATE)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        return wb
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作表1"
    headers = {
        1: "已回覆",
        3: "老師",
        4: "班別",
        5: "卷目",
        6: "學生編號",
        7: "班別",
        8: "學號",
        9: "姓名",
        10: "性別",
        11: "AFL",
        12: "平時分",
        13: "考試分",
        14: "AFL",
        15: "高水平",
        16: "考",
    }
    for col, text in headers.items():
        ws.cell(1, col, text)
    return wb


def _write_rows(ws, rows: list[dict]) -> None:
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 2, r["seq"])
        ws.cell(i, 3, r["staff"])
        ws.cell(i, 4, r["aux"])
        ws.cell(i, 5, r["paper"])
        ws.cell(i, 6, r["idStudent"])
        ws.cell(i, 7, r["class"])
        ws.cell(i, 8, r["num"])
        ws.cell(i, 9, r["name"])
        ws.cell(i, 10, r["gender"])
        ws.cell(i, 11, r["c11"])
        ws.cell(i, 12, r["c12"])
        ws.cell(i, 13, r["c13"])
        if r["c15"] is not None:
            ws.cell(i, 15, r["c15"])
        if r["c16"] is not None:
            ws.cell(i, 16, r["c16"])
        if r["c18"] is not None:
            ws.cell(i, 18, r["c18"])
        if r["c19"] is not None:
            ws.cell(i, 19, r["c19"])
        ws.cell(i, 20, r["c20"])
        ws.cell(i, 21, r["msg"])


def fetch_rows() -> list[dict]:
    cn = pyodbc.connect(CONN)
    cur = cn.cursor()

    score_rows: list[ScoreRow] = []
    for r in cur.execute(SCORE_SQL):
        score_rows.append(
            ScoreRow(
                idStaff=r.idStaff,
                auxiliaryClass=r.auxiliaryClass,
                idPaper=r.idPaper,
                idStudent=r.idStudent,
                class_=getattr(r, "class"),
                numberClass=r.numberClass,
                nameChinese=r.nameChinese,
                gender=r.gender,
                score_test_2=r.score_test_2,
                score_regular_2=r.score_regular_2,
                score_exam_2=r.score_exam_2,
                fm_test=r.fm_test,
                fm_regular=r.fm_regular,
                fm_exam=r.fm_exam,
            )
        )

    grade_rows = []
    for r in cur.execute(GRADE_SQL):
        grade_rows.append(
            {
                "idStaff": r.idStaff,
                "auxiliaryClass": r.auxiliaryClass,
                "idPaper": r.idPaper,
                "idStudent": r.idStudent,
                "class": getattr(r, "class"),
                "numberClass": r.numberClass,
                "nameChinese": r.nameChinese,
                "gender": r.gender,
            }
        )
    cn.close()

    out = _expand_score_rows(score_rows) + _expand_grade_rows(grade_rows)
    out.sort(
        key=lambda x: (
            x["staff"],
            x["aux"],
            x["paper"],
            x["class"],
            x["num"],
        )
    )
    _assign_seq(out)
    return out


def main() -> None:
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUT
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows = fetch_rows()
    wb = _load_workbook()
    _write_rows(wb.active, rows)
    wb.save(out_path)

    print(f"Wrote {len(rows)} rows -> {out_path}")
    staff_counts: dict[str, int] = {}
    for r in rows:
        staff_counts[r["staff"]] = staff_counts.get(r["staff"], 0) + 1
    print(f"Teachers: {len(staff_counts)}")
    for staff, n in sorted(staff_counts.items(), key=lambda x: (-x[1], x[0]))[:15]:
        print(f"  {staff}: {n}")


if __name__ == "__main__":
    main()
