#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate A_Done_SQL_Term2_DisciplineRecord.xls by copying 24-25 template and updating cells.

Preserves colleague Excel format / formulas; only StudentList + right-side PY data (M:AA) change.

See Administrative/CHW/student-report/guides/11_遲缺與考試缺席_Excel轉SQL.md
"""

from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    import pyodbc
    import win32com.client as win32
except ImportError:
    print("Install: pip install openpyxl pyodbc pywin32", file=sys.stderr)
    raise

from generate_term2_discipline_absent_sql import (  # noqa: E402
    DEFAULT_DISCIPLINE_XLSX,
    DISCIPLINE_DATA_START,
)
from _mssql_conn import connection_string  # noqa: E402

DEFAULT_TEMPLATE = Path(
    r"T:\24-25\ITAdmin_13_StudentReport\Datafile\24_25_Term2\A_Done_SQL_Term2_DisciplineRecord.xls"
)
DEFAULT_OUT = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\A_Done_SQL_Term2_DisciplineRecord.xls"
)

CONN = connection_string()

STUDENT_SQL = """
SELECT class + CAST(numberClass AS varchar), class, numberClass, nameChinese, idStudent
FROM dbo.tblStudent
WHERE LEFT(class, 1) IN ('1','2','3','4','5')
ORDER BY class, numberClass
"""

DATA_START_ROW = 4
# Template right block M:AA — maps to 25-26 PY source column indices (0-based)
RIGHT_SRC_COLS = [
    1,   # M 學號
    0,   # N 班別
    5,   # O 升/留/離
    3,   # P 英文姓名
    2,   # Q 中文姓名
    4,   # R 性別
    6,   # S 更新計劃取消缺點
    7,   # T 共記缺點(已包括欠交功課)
    8,   # U 因紀律/遲到
    9,   # V 因功課
    10,  # W 因欠交功課而被警告(Y=有)
    11,  # X 因紀律/遲到 (2)
    12,  # Y 因功課 (2)
    15,  # Z 遲到下學期
    16,  # AA 缺席下學期
]
RIGHT_FIRST_COL = 13   # M
RIGHT_LAST_COL = 27    # AA


def read_source_raw_rows(xlsx_path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if "缺點" in name or "20260622" in name:
            sheet_name = name
            break
    ws = wb[sheet_name or wb.sheetnames[0]]
    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < DISCIPLINE_DATA_START:
            continue
        if not row[0] or not row[17]:
            continue
        rows.append(tuple(row))
    wb.close()
    return rows


def fetch_student_list() -> list[tuple]:
    cn = pyodbc.connect(CONN)
    try:
        cur = cn.cursor()
        cur.execute(STUDENT_SQL)
        return [tuple(r) for r in cur.fetchall()]
    finally:
        cn.close()


def excel_value(value):
    if value is None or value == "":
        return None
    return value


def ensure_data_rows(ws, n_students: int, current_last_row: int) -> int:
    """Extend template data rows (copy row 4 formulas/format) or trim extras."""
    last_needed = DATA_START_ROW + n_students - 1
    if last_needed > current_last_row:
        template_row = ws.Range(f"{DATA_START_ROW}:{DATA_START_ROW}")
        template_row.Copy()
        ws.Range(f"{current_last_row + 1}:{last_needed}").Insert()
        ws.Application.CutCopyMode = False
        current_last_row = last_needed
    elif last_needed < current_last_row:
        ws.Range(
            ws.Cells(last_needed + 1, RIGHT_FIRST_COL),
            ws.Cells(current_last_row, RIGHT_LAST_COL),
        ).ClearContents()
        current_last_row = last_needed
    return current_last_row


def populate_excel(
    template_path: Path,
    out_path: Path,
    source_path: Path,
) -> int:
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    raw_rows = read_source_raw_rows(source_path)
    students = fetch_student_list()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, out_path)

    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = None
    try:
        wb = xl.Workbooks.Open(str(out_path.resolve()))

        ws_sl = wb.Sheets("StudentList")
        old_sl_rows = ws_sl.UsedRange.Rows.Count
        for i, row in enumerate(students, start=1):
            for j, val in enumerate(row, start=1):
                ws_sl.Cells(i, j).Value = val
        if len(students) < old_sl_rows:
            ws_sl.Range(
                ws_sl.Cells(len(students) + 1, 1),
                ws_sl.Cells(old_sl_rows, 5),
            ).ClearContents()

        ws = wb.Sheets("S1-5_DisciplineRecord")
        current_last = ws.UsedRange.Rows.Count
        current_last = ensure_data_rows(ws, len(raw_rows), current_last)

        for offset, raw in enumerate(raw_rows):
            r = DATA_START_ROW + offset
            for j, src_idx in enumerate(RIGHT_SRC_COLS):
                col = RIGHT_FIRST_COL + j
                val = raw[src_idx] if src_idx < len(raw) else None
                ws.Cells(r, col).Value = excel_value(val)

        wb.Save()
        return len(raw_rows)
    finally:
        if wb is not None:
            wb.Close(SaveChanges=True)
        xl.Quit()


def main() -> int:
    template_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_TEMPLATE
    source_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_DISCIPLINE_XLSX
    out_path = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_OUT

    if not source_path.exists():
        print(f"Missing source file: {source_path}", file=sys.stderr)
        return 1

    n = populate_excel(template_path, out_path, source_path)
    print(f"Wrote {out_path} ({n} students, {datetime.now().isoformat(timespec='seconds')})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
