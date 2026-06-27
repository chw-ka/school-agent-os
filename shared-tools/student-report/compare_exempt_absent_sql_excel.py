#!/usr/bin/env python3
"""Compare A_Done_SQL_Term2_Exempt_Absent.xlsx vs exam absent SQL file."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

SQL_PATH = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\SQL"
    r"\02_Update_tblStudentPaperScore_exam_absent_term2.sql"
)
XLSX_PATH = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2"
    r"\A_Done_SQL_Term2_Exempt_Absent.xlsx"
)

UPDATE_RE = re.compile(
    r"UPDATE tblStudentPaperScore SET score_exam_2 = 0, "
    r"(flgAbsent_2|flgIgnore_2) = 1 "
    r"WHERE idStudent = (\d+) AND idPaper = '([^']+)';",
    re.I,
)


def load_sql_updates(path: Path) -> list[tuple[int, str, str]]:
    text = path.read_text(encoding="utf-8")
    rows = []
    for m in UPDATE_RE.finditer(text):
        flag, sid, paper = m.group(1), int(m.group(2)), m.group(3)
        kind = "absent" if flag.lower() == "flgabsent_2" else "exempt"
        rows.append((sid, paper, kind))
    return rows


def load_excel_updates(path: Path) -> list[tuple[int | None, str, str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    main = wb[wb.sheetnames[0]]
    sl_ws = wb["StudentList"]
    lookup: dict[str, int] = {}
    for r in range(1, sl_ws.max_row + 1):
        k = sl_ws.cell(r, 1).value
        if k:
            lookup[str(k)] = int(sl_ws.cell(r, 2).value)

    rows = []
    for r in range(3, main.max_row + 1):
        paper = main.cell(r, 7).value
        if not paper:
            continue
        cls = main.cell(r, 5).value
        num = main.cell(r, 6).value
        key = f"{cls}{int(float(num))}"
        sid = lookup.get(key)
        if sid is None and main.cell(r, 3).value:
            sid = int(main.cell(r, 3).value)
        absent = main.cell(r, 8).value == 1
        exempt = main.cell(r, 9).value == 1
        kind = "exempt" if exempt else ("absent" if absent else "unknown")
        rows.append((sid, str(paper), kind, key))
    wb.close()
    return rows


def main() -> int:
    sql_rows = load_sql_updates(SQL_PATH)
    excel_rows = load_excel_updates(XLSX_PATH)

    sql_set = set(sql_rows)
    excel_with_sid = [(s, p, k) for s, p, k, _ in excel_rows if s]
    excel_set = set(excel_with_sid)
    excel_missing = [x for x in excel_rows if not x[0]]

    only_sql = sorted(sql_set - excel_set)
    only_excel = sorted(excel_set - sql_set)

    sql_by_sp = {(s, p): k for s, p, k in sql_set}
    excel_by_sp = {(s, p): k for s, p, k in excel_set}
    flag_mismatch = [
        (sp, sql_by_sp[sp], excel_by_sp[sp])
        for sp in set(sql_by_sp) & set(excel_by_sp)
        if sql_by_sp[sp] != excel_by_sp[sp]
    ]

    print(f"SQL file: {len(sql_rows)} updates ({len(sql_set)} unique)")
    print(f"Excel: {len(excel_rows)} rows ({len(excel_set)} unique with idStudent)")
    print(f"Excel missing idStudent: {len(excel_missing)}")
    print()

    if only_sql:
        print(f"Only in SQL ({len(only_sql)}):")
        for x in only_sql:
            print(f"  {x}")
    if only_excel:
        print(f"Only in Excel ({len(only_excel)}):")
        for x in only_excel:
            print(f"  {x}")
    if flag_mismatch:
        print(f"Flag mismatch ({len(flag_mismatch)}):")
        for x in flag_mismatch:
            print(f"  {x}")

    consistent = not only_sql and not only_excel and not flag_mismatch and not excel_missing
    print()
    print("RESULT:", "FULLY CONSISTENT" if consistent else "NOT FULLY CONSISTENT")
    if not consistent:
        print(f"  SQL only: {len(only_sql)}")
        print(f"  Excel only: {len(only_excel)}")
        print(f"  Flag mismatches: {len(flag_mismatch)}")
        print(f"  Excel missing idStudent: {len(excel_missing)}")

    # Breakdown by kind
    from collections import Counter
    print()
    print("SQL by kind:", dict(Counter(k for _, _, k in sql_set)))
    print("Excel by kind:", dict(Counter(k for _, _, k in excel_set)))

    sql_sp = Counter((s, p) for s, p, _ in sql_rows)
    dups = [(k, v) for k, v in sql_sp.items() if v > 1]
    if dups:
        print()
        print(f"SQL duplicate idStudent+idPaper pairs: {len(dups)} ({sum(v - 1 for _, v in dups)} extra lines)")
        for (sid, paper), v in sorted(dups, key=lambda x: -x[1])[:10]:
            print(f"  idStudent={sid} idPaper={paper} x{v}")

    return 0 if consistent else 1


if __name__ == "__main__":
    raise SystemExit(main())
