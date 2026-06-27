from excel_utils import mark_excel, parse_formula, parse_range
from openpyxl.formula.translate import Translator


def marking_function(workbook):
    comments = ""
    submarks = [0, 0, 0, 0, 0, 0]

    ws = workbook.active

    for i in range(3, 22):
        v = str(ws['E'+str(i)].value)
        if v == "=SUM(B"+str(i)+":D"+str(i)+")":
            submarks[0] += 2
        elif v == "=SUM(B"+str(i)+",C"+str(i)+",D"+str(i)+")" or v == "=SUM(B"+str(i)+"+C"+str(i)+"+D"+str(i)+")":
            submarks[0] += 1
            comments += "E"+str(i)+": Use of SUM function with range\n"
        elif v == "=SUM(A"+str(i)+":D"+str(i)+")":
            submarks[0] += 1
            comments += "E"+str(i)+": Incorrect Range\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"
        

        v = str(ws['F'+str(i)].value)
        if v == "=AVERAGE(B"+str(i)+":D"+str(i)+")":
            submarks[1] += 2
        elif v == "=AVERAGE(A"+str(i)+":D"+str(i)+")":
            submarks[1] += 1
            comments += "F"+str(i)+": Incorrect Range\n"
        else:
            comments += "F"+str(i)+": Wrong formula\n"

        v = str(ws['G'+str(i)].value).replace(" ", "")
        if v == "=_xlfn.RANK.EQ(F"+str(i)+",F$3:F$21,0)" or v == "=_xlfn.RANK.EQ(F"+str(i)+",$F$3:$F$21,0)" \
                or v == "=_xlfn.RANK.EQ(E"+str(i)+",E$3:E$21,0)" or v == "=_xlfn.RANK.EQ(E"+str(i)+",$E$3:$E$21,0)" \
                or v == "=_xlfn.RANK.EQ(F"+str(i)+",F$3:F$21)" or v == "=_xlfn.RANK.EQ(F"+str(i)+",$F$3:$F$21)" \
                or v == "=_xlfn.RANK.EQ(E"+str(i)+",E$3:E$21)" or v == "=_xlfn.RANK.EQ(E"+str(i)+",$E$3:$E$21)" \
                or v == "=RANK(F"+str(i)+",F$3:F$21,0)" or v == "=RANK(F"+str(i)+",$F$3:$F$21,0)" or v == "=RANK(F"+str(i)+",F$3:F$21)" \
                or v == "=_xlfn.RANK.EQ(E"+str(i)+",$E$3:$E$21,)" or v == "=_xlfn.RANK.EQ(F"+str(i)+",$F$1:$F$21)" \
                or v == "=_xlfn.RANK.EQ($F"+str(i)+",F$3:F$21)" or v == "=_xlfn.RANK.EQ(F"+str(i)+",$F$3:$F$21,)":
            submarks[2] += 2
        elif "=_xlfn.RANK.EQ(" in v or "=(_xlfn.RANK.EQ(" in v:
            submarks[2] += 1
            comments += "E"+str(i)+": Incorrect Range\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"

    for i in ["B", "C", "D"]:
        v = str(ws[i+'22'].value)
        if v == "=MAX("+i+"3:"+i+"21)":
            submarks[3] += 2
        elif "=MAX(" in v:
            submarks[3] += 1
            comments += i+'22: Incorrect Range\n'
        else:
            comments += i+'22: Wrong formula\n'

        v = str(ws[i+'23'].value)
        if v == "=MIN("+i+"3:"+i+"21)":
            submarks[4] += 2
        elif "=MIN(" in v:
            submarks[4] += 1
            comments += i+'23: Incorrect Range\n'
        else:
            comments += i+'23: Wrong formula\n'

        v = str(ws[i+'24'].value)
        if v == "=COUNTIF("+i+"3:"+i+"21,\">=50\")" or v == "=COUNTIF("+i+"2:"+i+"21,\">=50\")":
            submarks[5] += 2
        elif "=COUNTIF(" in v:
            submarks[5] += 1
            comments += i+'24: Incorrect Range\n'
        else:
            comments += i+'24: Wrong formula\n'

    fullmarks = [38, 38, 38, 6, 6, 6]
    marks = int(sum([m/f*100 for m, f in zip(submarks, fullmarks)]) / len(fullmarks))
    return marks, comments, submarks


assignment = "試算表功課01"
mark_excel(assignment, marking_function)
