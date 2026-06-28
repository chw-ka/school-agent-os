from excel_utils import mark_excel, parse_formula, parse_range
from openpyxl.formula.translate import Translator


def marking_function(workbook):
    comments = ""
    submarks = [0, 0, 0, 0, 0]

    # get worksheet
    ws = workbook["比賽資料"]
    for i in range(2, 128):
        v = str(ws['J'+str(i)].value).replace(" ", "")
        if (v == "=I"+str(i)+"/100+H"+str(i)+"+G"+str(i)+"*60" or v == "=G"+str(i)+"*60+H"+str(i)+"+I"+str(i)+"/100" or v == "=I"+str(i)+"/100+G"+str(i)+"*60+H"+str(i)+"" or v == "=SUM(I"+str(i)+"/100+H"+str(i)+"+G"+str(i)+"*60)" or v == "=+G"+str(i)+"*60+H"+str(i)+"+I"+str(i)+"/100"):
            submarks[0] += 1
        else:
            comments += "J"+str(i)+": Wrong formula\n"

    solutions = ['2D24', '2B23', '2C1', '2E17', '1B24', '1C19', '1A25', '2E12']
    ws = workbook["決賽 - 男子100米"]
    for i in range(2, 10):
        v = str(ws['A'+str(i)].value) + str(ws['B'+str(i)].value)
        if v in solutions[i-2]:
            submarks[1] += 1
        else:
            comments += "A"+str(i)+": Wrong answer\n"

    solutions = ['1A9', '1D3', '1B29', '1D25', '2D14', '2A8', '1D19', '1B14']
    ws = workbook["決賽 - 男子200米"]
    for i in range(2, 10):
        v = str(ws['A'+str(i)].value) + str(ws['B'+str(i)].value)
        if v in solutions[i-2]:
            submarks[2] += 1
        else:
            comments += "A"+str(i)+": Wrong answer\n"

    solutions = ['1C10', '1C21', '1C29', '2D18', '2E2', '2D2', '2E3', '1D6']
    ws = workbook["決賽 - 女子100米"]
    for i in range(2, 10):
        v = str(ws['A'+str(i)].value) + str(ws['B'+str(i)].value)
        if v in solutions[i-2]:
            submarks[3] += 1
        else:
            comments += "A"+str(i)+": Wrong answer\n"

    solutions = ['2A23', '2C20', '2D17', '2B20', '1A18', '2E15', '1C29', '2C24']
    ws = workbook["決賽 - 女子200米"]
    for i in range(2, 10):
        v = str(ws['A'+str(i)].value) + str(ws['B'+str(i)].value)
        if v in solutions[i-2]:
            submarks[4] += 1
        else:
            comments += "A"+str(i)+": Wrong answer\n"

    fullmarks = [126,8,8,8,8]
    marks = int(sum([m/f*100 for m, f in zip(submarks, fullmarks)]) / len(fullmarks))
    return marks, comments, submarks


assignment = "試算表功課03"
mark_excel(assignment, marking_function)
