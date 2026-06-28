from excel_utils import mark_excel, parse_formula, parse_range
from openpyxl.formula.translate import Translator


def marking_function(workbook):
    comments = ""
    submarks = [0, 0, 0, 0, 0, 0]

    # get worksheet
    ws = workbook["資料"]
    for i in range(2, 10):
        v = str(ws['D'+str(i)].value).replace(" ", "")
        if (v in ["=_xlfn.RANK.EQ(C"+str(i)+",C$2:C$9,1)", "=_xlfn.RANK.EQ(C"+str(i)+",C$2:C$9,10)",
                  "=+_xlfn.RANK.EQ(C"+str(i)+",C$2:C$9,1)"]):
            submarks[0] += 2
        elif "=_xlfn.RANK.EQ(" in v:
            submarks[0] += 1
            comments += "D"+str(i)+": Incorrect Range or order\n"
        else:
            comments += "D"+str(i)+": Wrong formula\n"

        v = str(ws['E'+str(i)].value).replace(" ", "")
        if (v in ["=VLOOKUP(D"+str(i)+",A$13:B$20,2)", "=VLOOKUP(D"+str(i)+",$A$13:$B$20,2)",
                  "=VLOOKUP(D"+str(i)+",A$13:B$20,2,FALSE)", "=VLOOKUP(D"+str(i)+",$A$13:$B$20,2,FALSE)",
                  "=VLOOKUP(D"+str(i)+",A$12:B$20,2)", "=VLOOKUP(D"+str(i)+",$A$13:B$20,2)",
                  "=VLOOKUP(D"+str(i)+",$A$12:$B$20,2)", "=VLOOKUP(D"+str(i)+",A$13:B$20,2)",
                  "=VLOOKUP(D"+str(i)+",A$3:B$20,2)", "=+VLOOKUP(D"+str(i)+",A$13:B$20,2)"]):
            submarks[1] += 2
        elif "VLOOKUP" in v:
            submarks[1] += 1
            comments += "E"+str(i)+": Incorrect Range or order\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"

        v = str(ws['G'+str(i)].value).replace(" ", "")
        if (v in ["=_xlfn.RANK.EQ(F"+str(i)+",F$2:F$9,0)", "=_xlfn.RANK.EQ(F"+str(i)+",F$2:F$9)",
                  "=+_xlfn.RANK.EQ(F"+str(i)+",F$2:F$9,0)", "=_xlfn.RANK.EQ(F"+str(i)+",F$2:F$9,)",
                  "=_xlfn.RANK.EQ(F"+str(i)+",$F$2:$F$9,0)", "=_xlfn.RANK.EQ(F"+str(i)+",F$2:F$9,)"]):
            submarks[2] += 2
        elif "=_xlfn.RANK.EQ(" in v:
            submarks[2] += 1
            comments += "D"+str(i)+": Incorrect Range or order\n"
        else:
            comments += "D"+str(i)+": Wrong formula\n"

        v = str(ws['H'+str(i)].value).replace(" ", "")
        if ("VLOOKUP" not in v):
            comments += "H"+str(i)+": Wrong formula\n"

        elif "G"+str(i)+"," in v and (",A$13:" in v or ",$A$13:" in v or ",A$12" in v or ",$A$12" in v) and ",2" in v:
            submarks[3] += 2
        elif "VLOOKUP" in v:
            submarks[3] += 1
            comments += "E"+str(i)+": Incorrect Range or order\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"

    for i in range(13, 17):
        v = str(ws['E'+str(i)].value)
        if (v in ["=SUMIF(B$2:B$9,D"+str(i)+",E$2:E$9)+SUMIF(B$2:B$9,D"+str(i)+",H$2:H$9)", "=SUMIF(B$2:B$9,D"+str(i)+",H$2:H$9)+SUMIF(B$2:B$9,D"+str(i)+",E$2:E$9)"]):
            submarks[4] += 2
        elif "SUMIF" in v:
            submarks[4] += 1
            comments += "E"+str(i)+": Incorrect Range or order\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"

        v = str(ws['F'+str(i)].value)
        if (v in ["=_xlfn.RANK.EQ(E"+str(i)+",E$13:E$16,0)", "=_xlfn.RANK.EQ(E"+str(i)+",E$13:E$16)", "=_xlfn.RANK.EQ(E"+str(i)+",E$13:E$16,)"]):
            submarks[5] += 2
        elif "=_xlfn.RANK.EQ(" in v:
            submarks[5] += 1
            comments += "F"+str(i)+": Incorrect Range or order\n"
        else:
            comments += "F"+str(i)+": Wrong formula\n"

    fullmarks = [16, 16, 16, 16, 8, 8]
    marks = int(sum([m/f*100 for m, f in zip(submarks, fullmarks)]) / len(fullmarks))
    return marks, comments, submarks


assignment = "試算表功課04"
mark_excel(assignment, marking_function)
