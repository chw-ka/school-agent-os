from excel_utils import mark_excel, parse_formula, parse_range
from openpyxl.formula.translate import Translator


def marking_function(workbook):
    comments = ""
    submarks = [0, 0, 0, 0, 0]

    # get worksheet 統計
    ws = workbook["統計"]

    for i in range(4, 12):
        v = str(ws['C'+str(i)].value)
        if v == "=COUNTIF(資料!E$2:G$128,B"+str(i)+")" or v == "=COUNTIF(資料!E$2:G$128,統計!B"+str(i)+")" \
                or v == "=COUNTIF(資料!D$2:G$128,統計!B"+str(i)+")" or v == "=COUNTIF(資料!D$2:G$128,B"+str(i)+")" \
                or v == "=COUNTIF(資料!$D$2:$G$128,B"+str(i)+")" or v == "=COUNTIF(資料!E$2:$G$128,統計!B"+str(i)+")" \
                or v == "=COUNTIF(資料!$D$2:$G$128,統計!B"+str(i)+")" or v == "=COUNTIF(資料!D$2:$G$128,B"+str(i)+")" \
                or v == "=COUNTIF(資料!A$2:H$128,B"+str(i)+")" or v == "=COUNTIF(資料!E$1:G$128,B"+str(i)+")" \
                or v == "=COUNTIFS(資料!E$2:G$128,B"+str(i)+")" or v == "=COUNTIFS(資料!D:D,\"MC\",資料!E:E,B"+str(i)+")+COUNTIFS(資料!D:D,\"MC\",資料!F:F,B"+str(i)+")+COUNTIFS(資料!D:D,\"MC\",資料!G:G,B"+str(i)+")":
            submarks[0] += 3
        elif v == "=COUNTIF(資料!D2:G128,B"+str(i)+")" or v == "=COUNTIF(資料!D2:G128,\"" + ws['B'+str(i)].value + "\")" \
                or v == "=COUNTIF(資料!E2:G128,B"+str(i)+")" or v == "=COUNTIF(資料!E2:G$128,B"+str(i)+")" \
                or v == "=COUNTIF(資料!D1:G128,B"+str(i)+")":
            submarks[0] += 2
            comments += "C"+str(i)+": No cell reference\n"
        elif "COUNTIF" in v:
            submarks[0] += 1
            comments += "C"+str(i)+": Incorrect Range\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"

        v = str(ws['D'+str(i)].value)
        if v == "=COUNTIF(資料!E$129:G$211,B"+str(i)+")" or v == "=COUNTIF(資料!E$129:G$211,統計!B"+str(i)+")" \
                or v == "=COUNTIF(資料!D$129:G$211,統計!B"+str(i)+")" or v == "=COUNTIF(資料!D$129:G$211,B"+str(i)+")" \
                or v == "=COUNTIF(資料!$D$129:$G$211,B"+str(i)+")" or v == "=COUNTIF(資料!E$129:$G$211,統計!B"+str(i)+")" \
                or v == "=COUNTIF(資料!$D$129:$G$211,統計!B"+str(i)+")" or v == "=COUNTIF(資料!D$129:$G$211,B"+str(i)+")" \
                or v == "=COUNTIF(資料!A$129:H$211,B"+str(i)+")" or v == "=COUNTIF(資料!E$129:G$211,B"+str(i)+")" \
                or v == "=COUNTIFS(資料!E$129:G$211,B"+str(i)+")" or v == "=COUNTIFS(資料!D:D,\"FC\",資料!E:E,B"+str(i)+")+COUNTIFS(資料!D:D,\"FC\",資料!F:F,B"+str(i)+")+COUNTIFS(資料!D:D,\"FC\",資料!G:G,B"+str(i)+")":
            submarks[1] += 3
        elif v == "=COUNTIF(資料!D129:G211,B"+str(i)+")" or v == "=COUNTIF(資料!D129:G211,\"" + ws['B'+str(i)].value + "\")" \
                or v == "=COUNTIF(資料!E129:G211,B"+str(i)+")" or v == "=COUNTIF(資料!E129:G$211,B"+str(i)+")" \
                or v == "=COUNTIF(資料!D129:G211,B"+str(i)+")" or v == "=COUNTIF(資料!E211:G211,統計!B"+str(i)+")" \
                or v == "=COUNTIF(資料!E129:G211,B"+str(i)+")":
            submarks[1] += 2
            comments += "C"+str(i)+": No cell reference\n"
        elif "COUNTIF" in v:
            submarks[1] += 1
            comments += "C"+str(i)+": Incorrect Range\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"

        v = str(ws['E'+str(i)].value)
        if v == "=SUM(C"+str(i)+":D"+str(i)+")" or v == "=SUM(C"+str(i)+",D"+str(i)+")" or v == "=SUM(C"+str(i)+",D"+str(i)+",)" or v == "=SUM(C"+str(i)+":D"+str(i)+",)":
            submarks[2] += 2
        elif "=SUM(" in v:
            submarks[2] += 1
            comments += "E"+str(i)+": Incorrect Range\n"
        else:
            comments += "E"+str(i)+": Wrong formula\n"


    for i in ["C", "D"]:
        v = str(ws[i+'12'].value)
        if v == "=SUM("+i+"4:"+i+"11)":
            submarks[3] += 2
        elif "=SUM(" in v:
            submarks[3] += 1
            comments += i+'12: Incorrect Range\n'
        else:
            comments += i+'12: Wrong formula\n'

    v = str(ws['E12'].value)
    if v == "=SUM(C12:D12)" or v == "=SUM(C12,D12)" or v == "=SUM(E4:E11)":
        submarks[4] += 2
    elif "=SUM(" in v:
        submarks[4] += 1
        comments += "E12: Incorrect Range\n"
    else:
        comments += "E12: Wrong formula\n"

    fullmarks = [24, 24, 16, 4, 2]
    marks = int(sum([m/f*100 for m, f in zip(submarks, fullmarks)]) / len(fullmarks))
    return marks, comments, submarks


assignment = "試算表功課02"
mark_excel(assignment, marking_function)
