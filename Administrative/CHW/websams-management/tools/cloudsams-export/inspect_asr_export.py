#!/usr/bin/env python3
"""Inspect a CloudSAMS ASR export xlsx/xls — column schema only, local use."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

def _inspect_xlsx(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(max_row=5, values_only=True)
        preview = [list(r) for r in rows]
        header = preview[0] if preview else []
        sheets[name] = {
            "columns": [str(c) if c is not None else "" for c in header],
            "preview_rows": len(preview),
            "col_count": len(header),
        }
    wb.close()
    return sheets


def _inspect_xls(path: Path) -> dict:
    import xlrd

    wb = xlrd.open_workbook(path)
    sheets = {}
    for i in range(wb.nsheets):
        ws = wb.sheet_by_index(i)
        preview = [ws.row_values(r) for r in range(min(5, ws.nrows))]
        header = preview[0] if preview else []
        sheets[ws.name] = {
            "columns": [str(c) if c is not None and c != "" else "" for c in header],
            "preview_rows": len(preview),
            "col_count": len(header),
        }
    return sheets


def inspect(path: Path) -> dict:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        sheets = _inspect_xls(path)
    elif suffix in (".xlsx", ".xlsm"):
        sheets = _inspect_xlsx(path)
    else:
        sys.exit(f"Unsupported format: {suffix} (expected .xls or .xlsx)")
    return {"file": str(path), "sheets": sheets}


def main() -> None:
    p = argparse.ArgumentParser(description="Inspect CloudSAMS ASR export column schema")
    p.add_argument("xlsx", type=Path, help="Decrypted export .xlsx from CloudSAMS")
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "_generation" / "asr-export-schema.json",
    )
    args = p.parse_args()
    if not args.xlsx.is_file():
        sys.exit(f"File not found: {args.xlsx}")
    result = inspect(args.xlsx)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"\nWrote {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()
