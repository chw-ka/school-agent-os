#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Add 上學期上課表現 (lesson_1) column to 體育分_2526_NY_女.xlsx."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

EXCEL = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\體育分_2526_NY_女.xlsx"
)
T_DATAFILE = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
)
LESSON_CSV = T_DATAFILE / "ped_female_lesson1_2526.csv"
SHEET = "Girls下學期"
COL_HEADER = "上學期上課表現"
INSERT_AT = 5  # column E, after 姓名 (D)


def load_lesson_map() -> dict[int, str]:
    mapping: dict[int, str] = {}
    with LESSON_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            mapping[int(row["idStudent"])] = row["lesson_1"].strip()
    return mapping


def main() -> None:
    lesson_map = load_lesson_map()
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]

    if ws.cell(1, INSERT_AT).value == COL_HEADER:
        print(f"Column '{COL_HEADER}' already at column {INSERT_AT}; updating values only.")
    else:
        ws.insert_cols(INSERT_AT)
        cell = ws.cell(1, INSERT_AT, COL_HEADER)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    filled = empty = 0
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is None:
            continue
        try:
            sid_int = int(float(sid))
        except (TypeError, ValueError):
            continue
        val = lesson_map.get(sid_int)
        ws.cell(r, INSERT_AT, val if val else None)
        if val:
            filled += 1
        else:
            empty += 1

    wb.save(EXCEL)
    print(f"Saved: {EXCEL}")
    print(f"Filled lesson_1: {filled}; no record: {empty}")


if __name__ == "__main__":
    main()
