#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create CM / manual subject awards Excel in T: Datafile folder."""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\25_26_Term2_科獎及特別獎項_CM確認.xlsx"
)

# idStudent verified against db25_26
STUDENTS = {
    ("1C", 7): (25037, "鍾尉汶"),
    ("2A", 8): (24023, "莊星月"),
    ("3C", 7): (23128, "鄒慧嵐"),
    ("4A", 4): (22064, "陳浠嵐"),
    ("5A", 21): (21083, "謝依桐"),
    ("1B", 3): (25016, "周雨鑫"),
    ("2D", 5): (24024, "周靖童"),
    ("3D", 32): (23116, "許晨曦"),
    ("1C", 13): (25059, "林韻晴"),
    ("2A", 10): (24028, "鍾喜兒"),
    ("3C", 13): (23062, "李星月"),
    ("4A", 5): (22095, "陳樂知"),
    ("5B", 2): (24127, "陳天怡"),
    ("1C", 2): (25002, "蔡仁博"),
    ("2B", 15): (24067, "文皓霆"),
    ("3A", 4): (23019, "陳敬威"),
    ("4D", 11): (22043, "賴兆鴻"),
    ("5B", 4): (20032, "陳金汛"),
    ("1B", 14): (25044, "何沛庭"),
    ("2B", 11): (24059, "林日熙"),
    ("3C", 9): (23036, "余姵瑤"),
    ("4B", 8): (22008, "郭曦煥"),
    ("5A", 25): (21053, "王欣彤"),
    ("1C", 8): (25042, "何咏軒"),
    ("2B", 6): (22039, "黃燦泓"),
    ("3C", 22): (23093, "宋延耀"),
    ("3D", 25): (25145, "曾愷明"),
    ("4C", 18): (22021, "劉沛然"),
    ("4C", 21): (21075, "吳瑋琪"),
    ("5B", 22): (21081, "曾曉琳"),
}

# (獎項名稱, 科目, 負責, 來源, 班別, 學號, 姓名, 21_22#, 備註)
AWARDS = [
    # MY — 音樂科科獎（去年 SpecialCases 同名）
    ("音樂科科獎", "MUS", "MY", "2026-06-25 CM", "1C", 7, "鍾尉汶", 35, ""),
    ("音樂科科獎", "MUS", "MY", "2026-06-25 CM", "2A", 8, "莊星月", 35, ""),
    ("音樂科科獎", "MUS", "MY", "2026-06-25 CM", "3C", 7, "鄒慧嵐", 35, ""),
    ("音樂科科獎", "MUS", "MY", "2026-06-25 CM", "4A", 4, "陳浠嵐", 35, ""),
    ("音樂科科獎", "MUS", "MY", "2026-06-25 CM", "5A", 21, "謝依桐", 35, ""),
    ("音樂才能傑出表現獎", "MUS", "MY", "2026-06-25 CM", "3C", 7, "鄒慧嵐", None, "CM 另報"),
    ("音樂才能傑出表現獎", "MUS", "MY", "2026-06-25 CM", "4A", 4, "陳浠嵐", None, "CM 另報"),
    # VA — 視藝科科獎（去年 SpecialCases 用「視藝科科獎」）
    ("視藝科科獎", "ART", "VA", "2026-06-25 CM", "1B", 3, "周雨鑫", 40, ""),
    ("視藝科科獎", "ART", "VA", "2026-06-25 CM", "2D", 5, "周靖童", 40, ""),
    ("視藝科科獎", "ART", "VA", "2026-06-25 CM", "3D", 32, "許晨曦", 40, ""),
    # EC — 聖經科科獎（去年成績表用名；每級一名）
    # CM「全年聖經科成績優異」+「中四級聖經科科獎」實為同一獎；4A5 只保留一條
    ("聖經科科獎", "BBS", "EC", "2026-06-25 CM", "1C", 13, "林韻晴", 34, ""),
    ("聖經科科獎", "BBS", "EC", "2026-06-25 CM", "2A", 10, "鍾喜兒", 34, ""),
    ("聖經科科獎", "BBS", "EC", "2026-06-25 CM", "3C", 13, "李星月", 34, ""),
    ("聖經科科獎", "BBS", "EC", "2026-06-25 CM", "4A", 5, "陳樂知", 34, ""),
    ("聖經科科獎", "BBS", "EC", "2026-06-25 CM", "5B", 2, "陳天怡", 34, ""),
    # ES — 男子體育科科獎（2526下學期男子體育科豁免及科獎名單給KACM.docx）
    ("體育科科獎", "PED", "ES", "docx 科獎名單", "1C", 2, "蔡仁博", 37, "男子"),
    ("體育科科獎", "PED", "ES", "docx 科獎名單", "2B", 15, "文皓霆", 37, "男子"),
    ("體育科科獎", "PED", "ES", "docx 科獎名單", "3A", 4, "陳敬威", 37, "男子"),
    ("體育科科獎", "PED", "ES", "docx 科獎名單", "4D", 11, "賴兆鴻", 37, "男子"),
    ("體育科科獎", "PED", "ES", "docx 科獎名單", "5B", 4, "陳金汛", 37, "男子"),
    # NY — 女子體育科科獎（體育分_2526_NY_女.xlsx 各級 TOTAL_MARK 最高）
    ("體育科科獎", "PED", "NY", "體育分_2526_NY_女.xlsx", "1B", 14, "何沛庭", 37, "女子"),
    ("體育科科獎", "PED", "NY", "體育分_2526_NY_女.xlsx", "2B", 11, "林日熙", 37, "女子"),
    ("體育科科獎", "PED", "NY", "體育分_2526_NY_女.xlsx", "3C", 9, "余姵瑤", 37, "女子"),
    ("體育科科獎", "PED", "NY", "體育分_2526_NY_女.xlsx", "4B", 8, "郭曦煥", 37, "女子"),
    ("體育科科獎", "PED", "NY", "體育分_2526_NY_女.xlsx", "5A", 25, "王欣彤", 37, "女子"),
]

MISSING = [
    (
        "運動員獎學金",
        "",
        "教務",
        "21_22 獎項一覽 #11；去年 SpecialCases 有",
        "待確認",
        11,
    ),
    (
        "鍾禮信先生數學獎學金",
        "MTH",
        "教務",
        "21_22 獎項一覽 #10；去年 SpecialCases 有",
        "待確認",
        10,
    ),
    (
        "郭永強校長獎學金",
        "",
        "教務",
        "21_22 獎項一覽 #4；去年 SpecialCases 有",
        "每級一名",
        4,
    ),
    (
        "宗教活動傑出服務獎",
        "",
        "教務",
        "21_22 獎項一覽 #12；去年 SpecialCases 有",
        "待確認",
        12,
    ),
    (
        "其他學科科獎（自動計算）",
        "",
        "IT/教務",
        "_獎項計算.sql 產生",
        "中/英/數/理科等",
        "19-33",
    ),
]

EXEMPT_PED = [
    # ES — 男子（2526下學期男子體育科豁免及科獎名單給KACM.docx）
    ("體育科下學期豁免", "PED", "ES", "docx 豁免名單", "1C", 8, "何咏軒", "M", "不寫科獎"),
    ("體育科下學期豁免", "PED", "ES", "docx 豁免名單", "2B", 6, "黃燦泓", "M", ""),
    ("體育科下學期豁免", "PED", "ES", "docx 豁免名單", "3C", 22, "宋延耀", "M", ""),
    # NY — 女子（體育分_2526_NY_女.xlsx → sheet「豁免名單」）
    ("體育科下學期豁免", "PED", "NY", "體育分_2526_NY_女.xlsx", "3D", 25, "曾愷明", "F", "不寫科獎"),
    ("體育科下學期豁免", "PED", "NY", "體育分_2526_NY_女.xlsx", "4C", 18, "劉沛然", "F", "不寫科獎"),
    ("體育科下學期豁免", "PED", "NY", "體育分_2526_NY_女.xlsx", "4C", 21, "吳瑋琪", "F", "不寫科獎"),
    ("體育科下學期豁免", "PED", "NY", "體育分_2526_NY_女.xlsx", "5B", 22, "曾曉琳", "F", "不寫科獎"),
]


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人手科獎及特別獎"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "#",
        "獎項名稱（成績表用）",
        "科目",
        "負責老師",
        "來源",
        "班別",
        "學號",
        "姓名",
        "idStudent",
        "姓名DB核對",
        "21_22獎項一覽#",
        "備註",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    for i, row in enumerate(AWARDS, 1):
        cls, num, name = row[4], row[5], row[6]
        sid, dbname = STUDENTS.get((cls, num), (None, ""))
        if dbname == name:
            name_ok = "OK"
        elif dbname:
            name_ok = f"DB:{dbname}"
        else:
            name_ok = "查無學生"
        ws.append([i] + list(row[:4]) + [cls, num, name, sid, name_ok, row[7], row[8]])

    widths = [4, 28, 8, 10, 22, 6, 6, 10, 10, 12, 12, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws_ex = wb.create_sheet("體育豁免_參考")
    ws_ex.append(["下學期", "性別", "班別", "學號", "姓名", "idStudent", "來源", "來源檔", "備註"])
    for row in EXEMPT_PED:
        cls, num, name = row[4], row[5], row[6]
        sid, _ = STUDENTS.get((cls, num), (None, ""))
        ws_ex.append([2, row[7], cls, num, name, sid, row[2], row[3], row[8] or None])

    ws2 = wb.create_sheet("可能漏報_待跟進")
    h2 = ["狀態", "獎項", "科目", "負責", "說明", "備註", "21_22#"]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        cell = ws2.cell(1, c)
        cell.fill = PatternFill("solid", fgColor="ED7D31")
        cell.font = Font(bold=True, color="FFFFFF")
    for m in MISSING:
        ws2.append(["待確認"] + list(m))
    for i, w in enumerate([10, 28, 8, 8, 40, 30, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet("說明")
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for line in [
        "25-26 下學期：人手科獎及特別獎項（跟去年 SpecialCases 獎項名稱）",
        f"更新：{ts}",
        "",
        "聖經：去年成績表用「聖經科科獎」每級一名。CM 報「全年聖經科成績優異」",
        "及「中四級聖經科科獎」係同一獎；4A5 陳樂知只保留一條，唔自創獎項名。",
        "",
        "體育男子科獎：2526下學期男子體育科豁免及科獎名單給KACM.docx（ES）",
        "女子體育科科獎：體育分_2526_NY_女.xlsx（NY），每級一名",
        "體育豁免：男子 ES docx 3 人；女子 NY xlsx「豁免名單」4 人 — 見「體育豁免_參考」",
        "",
        "視藝：去年 SpecialCases 用「視藝科科獎」（唔用「視覺藝術科科獎」）",
    ]:
        ws3.append([line])
    ws3.column_dimensions["A"].width = 100

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Award rows: {len(AWARDS)}")


if __name__ == "__main__":
    main()
