"""Update 體育豁免_參考 sheet in CM awards workbook."""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

XLSX = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\25_26_Term2_科獎及特別獎項_CM確認.xlsx"
)
T_DATAFILE = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
)
CSV = T_DATAFILE / "ped_exempt_2526_term2.csv"


def main() -> None:
    rows = list(csv.DictReader(CSV.open(encoding="utf-8-sig")))
    wb = openpyxl.load_workbook(XLSX)
    if "體育豁免_參考" in wb.sheetnames:
        del wb["體育豁免_參考"]
    ws = wb.create_sheet("體育豁免_參考")
    headers = ["下學期", "性別", "班別", "學號", "姓名", "idStudent", "來源", "來源檔", "備註"]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r["term"],
                r["gender"],
                r["class"],
                int(r["numberClass"]),
                r["nameChinese"],
                int(r["idStudent"]),
                r["source"],
                r["source_file"],
                r["remark"] or None,
            ]
        )
    for i, w in enumerate([8, 6, 6, 6, 10, 10, 8, 42, 14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(XLSX)
    print(f"Updated {XLSX} 體育豁免_參考 ({len(rows)} rows)")


if __name__ == "__main__":
    main()
