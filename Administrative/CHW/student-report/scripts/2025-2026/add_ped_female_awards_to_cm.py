"""Add 5 female PED subject awards to CM confirmation workbook."""
from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(
    r"T:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\25_26_Term2_科獎及特別獎項_CM確認.xlsx"
)

# (班別, 學號, 姓名, idStudent) — top female PED per form by TOTAL_MARK
FEMALE_PED_AWARDS = [
    ("1B", 14, "何沛庭", 25044),
    ("2B", 11, "林日熙", 24059),
    ("3C", 9, "余姵瑤", 23036),
    ("4B", 8, "郭曦煥", 22008),
    ("5A", 25, "王欣彤", 21053),
]


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["人手科獎及特別獎"]

    existing_ped_female = {
        (ws.cell(r, 5).value, ws.cell(r, 6).value)
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 2).value == "體育科科獎" and ws.cell(r, 11).value == "女子"
    }

    next_num = ws.max_row
    template_row = ws.max_row
    for cls, num, name, sid in FEMALE_PED_AWARDS:
        if (cls, num) in existing_ped_female:
            continue
        next_num += 1
        copy_row_style(ws, template_row, next_num, ws.max_column)
        ws.cell(next_num, 1, next_num - 1)
        ws.cell(next_num, 2, "體育科科獎")
        ws.cell(next_num, 3, "PED")
        ws.cell(next_num, 4, "NY")
        ws.cell(next_num, 5, cls)
        ws.cell(next_num, 6, num)
        ws.cell(next_num, 7, name)
        ws.cell(next_num, 8, sid)
        ws.cell(next_num, 9, "OK")
        ws.cell(next_num, 10, 37)
        ws.cell(next_num, 11, "女子")

    ws2 = wb["可能漏報_待跟進"]
    for r in range(2, ws2.max_row + 1):
        if ws2.cell(r, 2).value == "體育科學科獎-女同學":
            ws2.cell(r, 1, "已報")
            ws2.cell(r, 5, "體育分_2526_NY_女.xlsx 各級 TOTAL_MARK 最高；2026-06-26 加入 CM 確認表")
            ws2.cell(r, 6, "每級一名（5人）")

    ws3 = wb["說明"]
    for r in range(1, ws3.max_row + 1):
        if ws3.cell(r, 1).value and str(ws3.cell(r, 1).value).startswith("女子："):
            ws3.cell(r, 1, "女子體育科科獎：體育分_2526_NY_女.xlsx（NY），每級一名，已加入人手科獎表。")
    ws3.cell(ws3.max_row + 1, 1, f"女子體育科科獎更新：{datetime.now():%Y-%m-%d %H:%M}")

    wb.save(XLSX)
    print(f"Saved: {XLSX}")
    print("Added female PED awards:")
    for cls, num, name, sid in FEMALE_PED_AWARDS:
        print(f"  {cls} {num:02d} {name} ({sid})")


if __name__ == "__main__":
    main()
