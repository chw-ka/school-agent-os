"""Inspect and update CM awards xlsx."""
import sys
from pathlib import Path

import openpyxl
import pandas as pd

path = Path(
    r"t:\25-26\ITAdmin_13_StudentReport\Datafile\25_26_Term2\Datafile"
    r"\25_26_Term2_科獎及特別獎項_CM確認.xlsx"
)
wb = openpyxl.load_workbook(path)
ws = wb[wb.sheetnames[0]]
print("Sheet:", ws.title, "max_row", ws.max_row)
for r in range(1, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 12)]
    print(r, row)
